import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

def extract_hash_links(url):
    """Extract all hash (#...) links from a given URL."""
    try:
        response = requests.get(url)
        response.raise_for_status()
    except Exception as e:
        print(f"❌ Failed to fetch {url}: {e}")
        return []

    soup = BeautifulSoup(response.text, "html.parser")
    links = soup.find_all("a", href=True)

    results = []
    for link in links:
        href = link["href"].strip()
        if href.startswith("#"):
            link_text = link.get_text(strip=True) or "(no text)"
            results.append({
                "Page URL": url,
                "Link Text": link_text,
                "Hash Link": href
            })
    return results


def create_excel_report(all_links, output_file="Hash_Link_Report.xlsx"):
    """Generate a formatted Excel report from extracted links."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hash Links"

    # Header style
    header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["#", "Page URL", "Link Text", "Hash Link"]
    ws.append(headers)

    # Apply header style
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for i, link in enumerate(all_links, start=1):
        ws.append([i, link["Page URL"], link["Link Text"], link["Hash Link"]])

    # Adjust column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25

    # Add borders and alignments
    for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Create Summary Sheet
    summary = wb.create_sheet("Summary")
    summary["A1"] = "Hash Link Report"
    summary["A1"].font = Font(size=16, bold=True, color="1E88E5")

    summary["A3"] = "Generated On:"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    summary["A4"] = "Total Links Found:"
    summary["B4"] = len(all_links)

    summary.column_dimensions["A"].width = 25
    summary.column_dimensions["B"].width = 50

    wb.save(output_file)
    print(f"✅ Professional Excel report generated: {output_file}")


if __name__ == "__main__":
    # URLs 
    urls = [
        "https://mrkararelal.com/",
        "https://mrkararelal.com/shop/",
        "https://mrkararelal.com/about-us/",
        "https://mrkararelal.com/contact-us/"
    ]

    all_links = []
    for url in urls:
        all_links.extend(extract_hash_links(url))

    if all_links:
        create_excel_report(all_links)
    else:
        print("⚠️ No hash links found on the provided pages.")
